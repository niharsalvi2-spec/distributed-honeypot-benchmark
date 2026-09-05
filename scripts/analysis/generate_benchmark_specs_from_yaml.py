"""
Automated Generator: Compiles benchmark_specs/*.xlsx from machine-readable YAML configs.
Source of truth: configs/ (YAML/JSON)
Generated Artifacts: benchmark_specs/*.xlsx
"""
import os
import sys
import yaml
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

def create_styled_sheet(ws, title, headers, rows):
    ws.title = title
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

def generate_repository_register():
    cfg_file = os.path.join(ROOT, "configs", "repositories.yaml")
    with open(cfg_file, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f).get("repositories", {})
    
    headers = ["ID", "Name", "Official URL", "Tier", "Version", "Distributed Support", "Audit Status"]
    rows = []
    idx = 1
    for name, info in data.items():
        rows.append([
            f"R{idx:02d}",
            name.upper(),
            f"https://github.com/{name}/{name}",
            info.get("tier", "primary").capitalize(),
            info.get("version", "latest"),
            "Distributed Forwarder" if name == "mhn" else "Single Node / Decoy",
            "AUDITED & EXTRACTED"
        ])
        idx += 1
    
    wb = openpyxl.Workbook()
    create_styled_sheet(wb.active, "Repositories", headers, rows)
    out_path = os.path.join(ROOT, "benchmark_specs", "repository_register.xlsx")
    wb.save(out_path)
    print(f"Compiled: {out_path}")

def generate_experiment_matrix():
    exp_files = sorted(glob.glob(os.path.join(ROOT, "configs", "experiments", "E*.yaml")))
    headers = ["Exp ID", "Description", "Iterations", "Timeout (s)", "Metric Logging", "Target Phase"]
    rows = []
    for ef in exp_files:
        with open(ef, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        rows.append([
            cfg.get("experiment_id", os.path.basename(ef).replace(".yaml", "")),
            cfg.get("description", "Automated Experiment"),
            cfg.get("iterations", 10),
            cfg.get("timeout_seconds", 300),
            "ENABLED" if cfg.get("collect_metrics") else "DISABLED",
            "Phase 4 (Distributed Evaluation)"
        ])
    wb = openpyxl.Workbook()
    create_styled_sheet(wb.active, "Experiment Matrix", headers, rows)
    out_path = os.path.join(ROOT, "benchmark_specs", "experiment_matrix.xlsx")
    wb.save(out_path)
    print(f"Compiled: {out_path}")

def generate_metric_dictionary():
    cfg_file = os.path.join(ROOT, "configs", "metrics.yaml")
    with open(cfg_file, "r", encoding="utf-8") as f:
        metrics_cfg = yaml.safe_load(f).get("metrics", {})
    
    headers = ["Category", "Metric Identifier", "Unit", "Validation Target"]
    rows = []
    for cat, m_list in metrics_cfg.items():
        for m in m_list:
            unit = "Ratio [0-1]" if m in ["precision", "recall", "f1"] else ("EPS" if m == "throughput" else "ms / %")
            rows.append([cat.title(), m.upper(), unit, "Continuous Evaluation"])
            
    wb = openpyxl.Workbook()
    create_styled_sheet(wb.active, "Metrics", headers, rows)
    out_path = os.path.join(ROOT, "benchmark_specs", "metric_dictionary.xlsx")
    wb.save(out_path)
    print(f"Compiled: {out_path}")

def main():
    print("=== Compiling Benchmark Specifications from Machine-Readable YAML Configs ===")
    generate_repository_register()
    generate_experiment_matrix()
    generate_metric_dictionary()
    print("Compilation completed successfully.")

if __name__ == "__main__":
    main()
